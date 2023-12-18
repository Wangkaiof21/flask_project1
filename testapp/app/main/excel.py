#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2021/3/12 0:56
# @Author  : v_bkaiwang
# @File    : excel.py
# @Software: win10 Tensorflow1.13.1 python3.6.3
import os
import openpyxl


class Excel:
    def __init__(self, file: str):
        self.file_name = file
        if os.path.exists(file):
            self.wb = openpyxl.load_workbook(file)
        else:
            self.wb = openpyxl.Workbook()

    def sheet2dict(self):
        perf_total = self.wb['perf_total']
        rows = [row for row in perf_total.rows]
        headers = self.get_headers(rows)
        test_case_name_chk = []
        result = []
        for row in rows[1:]:
            row = [cell.value for cell in row]
            line_data = dict(zip(headers, row))
            if line_data['TestCaseName'] not in test_case_name_chk:
                line_data['result'] = self.subsheet2dict(line_data['Benchmark'])
                test_case_name_chk.append(line_data['TestCaseName'])
                result.append(line_data)
        return result

    def subsheet2dict(self, sheet_name):
        sheet = self.wb[sheet_name]
        rows = [row for row in sheet.rows]
        headers = self.get_headers(rows)
        result = []
        for row in rows[1:]:
            row = [cell.value for cell in row]
            line_data = dict(zip(headers, row))
            result.append(line_data)
        return result

    def get_headers(self, rows):
        headers = [cell.value for cell in rows[0]]
        for i in range(len(headers)):
            if "." in headers[i]:
                headers[i] = headers[i].replace('.', 'p')
        return headers

    def results2file(self, results: list) -> None:
        """
        将数据的数据转化成文件保存
        :param results:
        :return:
        """
        results = [i for i in results]
        benchmark_list = list(set(i['Benchmark'] for i in results))
        for benchmark in benchmark_list:
            self.wb.create_sheet(benchmark)
        del self.wb['Sheet']
        """ 初始化表格写入的单元坐标 """
        row, col = 1, 1
        """遍历所有数据"""
        for result in results:
            benchmark = result["Benchmark"]
            sheet = self.wb[benchmark]
            """生成拼接后的sub_results_"""
            sub_results_ = list()
            sub_results__list = result['result']
            del result['result']
            del result['Str3Compare']
            for sub_result in sub_results__list:
                temp = result.copy()
                temp.update(sub_result)
                sub_results_.append(temp)
            """根据sub_results处理表头"""
            sub_results_ = sub_results_[0]
            for key in sub_results_:
                sheet.cell(row, col).value = key
                col += 1
            col = 1
            row += 1
            """根据sub_results处理表数据"""
            for sub_result in sub_results_:
                for _, value in sub_result.items():
                    sheet.cell(row, col).value = value
                    col += 1
                row += 1
                col = 1
            row = 1
            """保存文件"""
            self.wb.save(self.file_name)

    def keys_to_template_file(self, sheet_name, keys) -> None:
        for workbook in self.wb:
            del self.wb[workbook.title]
        self.wb.create_sheet(sheet_name)
        sheet = self.wb[sheet_name]
        row, col = 1, 1
        for key in keys:
            sheet.cell(row, col).value = key
            col += 1
        self.wb.save(self.file_name)


if __name__ == '__main__':
    print(Excel('./upload/template/template_bw_mem.xlsx').sheet2dict())
